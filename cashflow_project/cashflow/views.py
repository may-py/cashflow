"""
views.py  —  Cash Flow dashboard + API + exports
  GET /cashflow/                   → HTML dashboard
  GET /cashflow/api/               → JSON
  GET /cashflow/export/excel/      → .xlsx
  GET /cashflow/export/pdf/        → .pdf
Query params (all endpoints):
  date_from    YYYY-MM-DD  (default: today)
  date_to      YYYY-MM-DD  (default: today + 90 days)
  company      1 | 2 | all (default: all)
  group_days   int         (default: 7, pivot granularity)
"""

from __future__ import annotations

import io
import logging
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET

from .odoo_client import OdooAPIError, fetch_cashflow_data
from .services import build_projection, CashflowProjection

logger = logging.getLogger(__name__)

COMPANIES = {1: 'KOB', 2: 'BTV'}


# ── Query-param helpers ───────────────────────────────────────────────────────

def _parse_params(request: HttpRequest):
    today = date.today()

    # Dates
    try:
        date_from = date.fromisoformat(request.GET.get('date_from', ''))
    except ValueError:
        date_from = today
    try:
        date_to = date.fromisoformat(request.GET.get('date_to', ''))
    except ValueError:
        date_to = today + timedelta(days=90)

    # Company filter
    company_param = request.GET.get('company', 'all')
    if company_param in ('1', '2'):
        company_ids = [int(company_param)]
    else:
        company_ids = [1, 2]

    # Pivot group size
    try:
        group_days = max(1, int(request.GET.get('group_days', 7)))
    except (ValueError, TypeError):
        group_days = 7

    # Opening balance (starting cash position)
    try:
        opening_balance = Decimal(request.GET.get('opening_balance', '0') or '0')
    except Exception:
        opening_balance = Decimal('0')

    return date_from, date_to, company_ids, group_days, company_param, opening_balance


def _get_projection(date_from, date_to, company_ids, opening_balance=Decimal('0')) -> CashflowProjection | None:
    try:
        raw = fetch_cashflow_data(date_from, date_to, company_ids)
        return build_projection(raw, date_from, date_to, opening_balance=opening_balance)
    except OdooAPIError as exc:
        logger.error('Odoo API error: %s', exc)
        return None
    except Exception as exc:
        logger.exception('Unexpected error: %s', exc)
        return None


# ── HTML Dashboard ────────────────────────────────────────────────────────────

@require_GET
@login_required
def dashboard(request: HttpRequest) -> HttpResponse:
    date_from, date_to, company_ids, group_days, company_param, opening_balance = _parse_params(request)
    projection = _get_projection(date_from, date_to, company_ids, opening_balance)
    pivot      = projection.pivot(group_days) if projection else []

    # Both charts now load async via API — no chart data embedded in HTML
    context = {
        'date_from':        date_from.isoformat(),
        'date_to':          date_to.isoformat(),
        'company':          company_param,
        'group_days':       group_days,
        'opening_balance':  opening_balance,
        'projection':       projection,
        'pivot':            pivot,
        'panels':           [('all', 'all'), ('ar', 'receivable'), ('ap', 'payable')],
        'error':            None if projection else 'Could not connect to Odoo. Check your settings.',
    }
    return render(request, 'cashflow/dashboard.html', context)


# ── JSON API ──────────────────────────────────────────────────────────────────

@require_GET
@login_required
def api_cashflow(request: HttpRequest) -> JsonResponse:
    date_from, date_to, company_ids, group_days, company_param, opening_balance = _parse_params(request)
    projection = _get_projection(date_from, date_to, company_ids, opening_balance)

    if projection is None:
        return JsonResponse({'error': 'Failed to fetch Odoo data'}, status=502)

    pivot = projection.pivot(group_days)

    return JsonResponse({
        'date_from':          date_from.isoformat(),
        'date_to':            date_to.isoformat(),
        'company':            company_param,
        'group_days':         group_days,
        'total_inflow_thb':   float(projection.total_inflow_thb),
        'total_outflow_thb':  float(projection.total_outflow_thb),
        'net_position_thb':   float(projection.net_position_thb),
        'pivot': [
            {
                'period':      p.period_label,
                'date_from':   p.date_from.isoformat(),
                'date_to':     p.date_to.isoformat(),
                'inflow_thb':  float(p.inflow_thb),
                'outflow_thb': float(p.outflow_thb),
                'net_thb':     float(p.net_thb),
                'running_thb': float(p.running_thb),
                'lines': [
                    {
                        'id':         l.line_id,
                        'ref':        l.entry_ref,
                        'partner':    l.partner,
                        'due_date':   l.due_date.isoformat() if l.due_date else None,
                        'amount_thb': float(l.amount_thb),
                        'currency':   l.currency,
                        'move_ref':   l.move_ref,
                        'type':       l.line_type,
                        'company':    l.company,
                        'company_id': l.company_id,
                    }
                    for l in p.lines
                ],
            }
            for p in pivot
        ],
    })


# ── Excel Export ──────────────────────────────────────────────────────────────

@require_GET
@login_required
def export_excel(request: HttpRequest) -> HttpResponse:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse('openpyxl not installed. Run: pip install openpyxl', status=500)

    date_from, date_to, company_ids, group_days, company_param, opening_balance = _parse_params(request)
    projection = _get_projection(date_from, date_to, company_ids, opening_balance)
    if projection is None:
        return HttpResponse('Failed to fetch Odoo data', status=502)

    pivot  = projection.pivot(group_days)
    wb     = openpyxl.Workbook()

    DARK   = '1A2B4C'
    GREEN  = '00C48C'
    RED    = 'FF5252'
    BLUE   = '4A90D9'
    LIGHT  = 'F0F4FA'

    def hf(size=10, bold=True, color='FFFFFF'):
        return Font(name='Calibri', size=size, bold=bold, color=color)
    def df(size=10, bold=False, color='1A2B4C'):
        return Font(name='Calibri', size=size, bold=bold, color=color)
    def fill(c):
        return PatternFill('solid', fgColor=c)
    def bdr():
        s = Side(style='thin', color='D0D8E8')
        return Border(left=s, right=s, top=s, bottom=s)
    def num(ws, row, col, val, fmt='#,##0.00', color=None):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = fmt
        c.font   = Font(name='Calibri', size=10, bold=bool(color),
                        color=color or '1A2B4C')
        c.border = bdr()
        c.alignment = Alignment(horizontal='right')
        return c

    company_label = ('KOB' if company_param == '1' else 'BTV' if company_param == '2' else 'KOB + BTV')

    # ── Sheet 1: Pivot Summary ────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Pivot (THB)'
    ws['A1'] = 'Cash Flow Projection — THB'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=DARK)
    ws['A2'] = (f'Period: {date_from:%d %b %Y} – {date_to:%d %b %Y}  |  '
                f'{company_label}  |  Group: {group_days} days')
    ws['A2'].font = Font(name='Calibri', size=10, color='5A6A8A')

    # KPI row
    for col, (lbl, val, c) in enumerate([
        ('Total Inflows (THB)',  float(projection.total_inflow_thb),  GREEN),
        ('Total Outflows (THB)', float(projection.total_outflow_thb), RED),
        ('Net Position (THB)',   float(projection.net_position_thb),  BLUE),
    ], 1):
        h = ws.cell(row=4, column=col, value=lbl)
        v = ws.cell(row=5, column=col, value=val)
        h.font = hf(color=c); h.fill = fill('EBF3FD')
        h.alignment = Alignment(horizontal='center')
        v.font = Font(name='Calibri', size=14, bold=True,
                      color=(GREEN if val >= 0 else RED))
        v.number_format = '#,##0.00'
        v.alignment = Alignment(horizontal='center')

    # Pivot table header
    hdrs = ['Period', 'Inflow (THB)', 'Outflow (THB)', 'Net (THB)', 'Running Balance (THB)']
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = hf(); c.fill = fill(DARK)
        c.alignment = Alignment(horizontal='center')

    for ri, p in enumerate(pivot, 8):
        ws.cell(row=ri, column=1, value=p.period_label).font = df()
        ws.cell(row=ri, column=1).border = bdr()
        num(ws, ri, 2, float(p.inflow_thb),  color=GREEN)
        num(ws, ri, 3, float(p.outflow_thb))
        num(ws, ri, 4, float(p.net_thb),
            color=(GREEN if p.net_thb >= 0 else RED))
        num(ws, ri, 5, float(p.running_thb),
            color=(GREEN if p.running_thb >= 0 else RED))
        if ri % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=ri, column=col).fill = fill(LIGHT)

    ws.column_dimensions['A'].width = 28
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # ── Sheet 2: AR Detail ────────────────────────────────────────────────────
    ws_ar = wb.create_sheet('Receivables (AR)')
    ar_hdrs = ['Period', 'Due Date', 'Company', 'Partner', 'Entry Ref',
               'Journal Entry', 'Orig Currency', 'Amount (THB)']
    for col, h in enumerate(ar_hdrs, 1):
        c = ws_ar.cell(row=1, column=col, value=h)
        c.font = hf(); c.fill = fill(GREEN)
        c.alignment = Alignment(horizontal='center')

    ri = 2
    for p in pivot:
        for line in p.lines:
            if line.line_type != 'receivable':
                continue
            ws_ar.cell(row=ri, column=1, value=p.period_label)
            ws_ar.cell(row=ri, column=2, value=line.due_date.strftime('%d %b %Y') if line.due_date else '')
            ws_ar.cell(row=ri, column=3, value=line.company)
            ws_ar.cell(row=ri, column=4, value=line.partner)
            ws_ar.cell(row=ri, column=5, value=line.entry_ref)
            ws_ar.cell(row=ri, column=6, value=line.move_ref)
            ws_ar.cell(row=ri, column=7, value=line.currency)
            num(ws_ar, ri, 8, float(line.amount_thb), color=GREEN)
            for col in range(1, 8):
                ws_ar.cell(row=ri, column=col).font   = df()
                ws_ar.cell(row=ri, column=col).border = bdr()
                if ri % 2 == 0:
                    ws_ar.cell(row=ri, column=col).fill = fill('E8FAF4')
            ri += 1

    for col, w in zip(range(1, 9), [22, 14, 20, 28, 22, 22, 12, 18]):
        ws_ar.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: AP Detail ────────────────────────────────────────────────────
    ws_ap = wb.create_sheet('Payables (AP)')
    for col, h in enumerate(ar_hdrs, 1):
        c = ws_ap.cell(row=1, column=col, value=h)
        c.font = hf(); c.fill = fill(RED)
        c.alignment = Alignment(horizontal='center')

    ri = 2
    for p in pivot:
        for line in p.lines:
            if line.line_type != 'payable':
                continue
            ws_ap.cell(row=ri, column=1, value=p.period_label)
            ws_ap.cell(row=ri, column=2, value=line.due_date.strftime('%d %b %Y') if line.due_date else '')
            ws_ap.cell(row=ri, column=3, value=line.company)
            ws_ap.cell(row=ri, column=4, value=line.partner)
            ws_ap.cell(row=ri, column=5, value=line.entry_ref)
            ws_ap.cell(row=ri, column=6, value=line.move_ref)
            ws_ap.cell(row=ri, column=7, value=line.currency)
            num(ws_ap, ri, 8, float(abs(line.amount_thb)), color=RED)
            for col in range(1, 8):
                ws_ap.cell(row=ri, column=col).font   = df()
                ws_ap.cell(row=ri, column=col).border = bdr()
                if ri % 2 == 0:
                    ws_ap.cell(row=ri, column=col).fill = fill('FFF0F0')
            ri += 1

    for col, w in zip(range(1, 9), [22, 14, 20, 28, 22, 22, 12, 18]):
        ws_ap.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname    = f'cashflow_{date_from}_{date_to}_{company_param}_{group_days}d.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ── PDF Export ────────────────────────────────────────────────────────────────

@require_GET
@login_required
def export_pdf(request: HttpRequest) -> HttpResponse:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        return HttpResponse('reportlab not installed. Run: pip install reportlab', status=500)

    date_from, date_to, company_ids, group_days, company_param, opening_balance = _parse_params(request)
    projection = _get_projection(date_from, date_to, company_ids, opening_balance)
    if projection is None:
        return HttpResponse('Failed to fetch Odoo data', status=502)

    pivot  = projection.pivot(group_days)
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=landscape(A4),
                               rightMargin=1.5*cm, leftMargin=1.5*cm,
                               topMargin=1.5*cm, bottomMargin=1.5*cm)

    C_DARK  = colors.HexColor('#1A2B4C')
    C_GREEN = colors.HexColor('#00C48C')
    C_RED   = colors.HexColor('#FF5252')
    C_BLUE  = colors.HexColor('#4A90D9')
    C_LIGHT = colors.HexColor('#F0F4FA')

    def ps(name, **kw):
        return ParagraphStyle(name, **kw)

    title_s = ps('T', fontName='Helvetica-Bold', fontSize=16, textColor=C_DARK, spaceAfter=4)
    sub_s   = ps('S', fontName='Helvetica', fontSize=9, textColor=colors.HexColor('#5A6A8A'), spaceAfter=10)

    company_label = ('KOB' if company_param == '1' else 'BTV' if company_param == '2' else 'KOB + BTV')
    elems = [
        Paragraph('Cash Flow Projection — THB', title_s),
        Paragraph(
            f'Period: {date_from:%d %b %Y} – {date_to:%d %b %Y}  |  '
            f'{company_label}  |  Group: {group_days} days',
            sub_s,
        ),
        HRFlowable(width='100%', thickness=2, color=C_DARK),
        Spacer(1, 0.3*cm),
    ]

    # KPI row
    kpi_data = [[
        Paragraph(f'<b>Total Inflows (THB)</b><br/>'
                  f'<font color="#00C48C" size="14">{float(projection.total_inflow_thb):,.2f}</font>',
                  ps('k', fontName='Helvetica', fontSize=10, alignment=1)),
        Paragraph(f'<b>Total Outflows (THB)</b><br/>'
                  f'<font color="#FF5252" size="14">{float(projection.total_outflow_thb):,.2f}</font>',
                  ps('k2', fontName='Helvetica', fontSize=10, alignment=1)),
        Paragraph(f'<b>Net Position (THB)</b><br/>'
                  f'<font color="{"#00C48C" if projection.net_position_thb >= 0 else "#FF5252"}" size="14">'
                  f'{float(projection.net_position_thb):,.2f}</font>',
                  ps('k3', fontName='Helvetica', fontSize=10, alignment=1)),
    ]]
    kt = Table(kpi_data, colWidths=[8*cm, 8*cm, 8*cm])
    kt.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), C_LIGHT),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#D0D8E8')),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    elems += [kt, Spacer(1, 0.4*cm)]

    # Pivot table
    pw   = landscape(A4)[0] - 3*cm
    cols = [pw*0.30, pw*0.175, pw*0.175, pw*0.175, pw*0.175]
    hdr  = ['Period', 'Inflow (THB)', 'Outflow (THB)', 'Net (THB)', 'Running (THB)']
    rows = [hdr]
    for p in pivot:
        net_c = '#00C48C' if p.net_thb >= 0 else '#FF5252'
        run_c = '#00C48C' if p.running_thb >= 0 else '#FF5252'
        rows.append([
            p.period_label,
            f'{float(p.inflow_thb):,.2f}',
            f'{float(abs(p.outflow_thb)):,.2f}',
            Paragraph(f'<font color="{net_c}"><b>{float(p.net_thb):,.2f}</b></font>',
                      ps('n', fontName='Helvetica', fontSize=9)),
            Paragraph(f'<font color="{run_c}"><b>{float(p.running_thb):,.2f}</b></font>',
                      ps('r', fontName='Helvetica', fontSize=9)),
        ])

    pt = Table(rows, colWidths=cols)
    pt.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0),  C_DARK),
        ('TEXTCOLOR',     (0,0), (-1,0),  colors.white),
        ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 9),
        ('ALIGN',         (1,0), (-1,-1), 'RIGHT'),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [colors.white, C_LIGHT]),
        ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#D0D8E8')),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    elems.append(pt)

    doc.build(elems)
    buf.seek(0)
    fname    = f'cashflow_{date_from}_{date_to}_{company_param}_{group_days}d.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response